# ExcelKeywords.py
from openpyxl import load_workbook

def read_excel_row(file_path, sheet_name, row_num):
    wb = load_workbook(filename=file_path, data_only=True)
    sheet = wb[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[row_num]]
    return dict(zip(headers, values))

def read_all_rows(file_path, sheet_name):
    wb = load_workbook(filename=file_path, data_only=True)
    sheet = wb[sheet_name]
    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data




*** Settings ***
Library           ExcelKeywords.py
Library           Collections

*** Variables ***
${EXCEL_FILE}     test_data.xlsx
${SHEET_NAME}     Sheet1

*** Test Cases ***
Verify Row 2 Matches System
    ${excel_row}=    Read Excel Row    ${EXCEL_FILE}    ${SHEET_NAME}    2
    Log    ${excel_row}
    ${actual_name}=    Get Actual Name
    Should Be Equal As Strings    ${actual_name}    ${excel_row}[Name]

Verify All Excel Rows
    ${rows}=    Read All Rows    ${EXCEL_FILE}    ${SHEET_NAME}
    :FOR    ${row}    IN    @{rows}
    \    Log    Verifying: ${row}
    \    ${actual}=    Get Actual Name By Name    ${row}[Name]
    \    Should Be Equal As Strings    ${actual}    ${row}[Name]

*** Keywords ***
Get Actual Name
    [Return]    John Doe

Get Actual Name By Name
    [Arguments]    ${expected_name}
    # ใส่ logic จริง เช่น ดึงจากระบบ หรือจำลอง
    [Return]    ${expected_name}
